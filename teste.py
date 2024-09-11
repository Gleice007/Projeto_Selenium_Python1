import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from time import sleep
import smtplib
import os
from email.message import EmailMessage
import re

class Scrappy:

    def iniciar(self):
        self.obter_email_usuario_senha()
        self.raspagem_de_dados()
        self.criar_planilha()
        self.enviar_email_cliente()

    def obter_email_usuario_senha(self):
        self.email = input("Digite o email para receber o relatorio de valores dos celulares!\n")
        self.email = self.email.lower()
        self.senha = input("Digite a sua senha: ")
        padrao = re.search(r'[a-zA-Z0-9_-]+@[a-zA-Z0-9]+\.[a-zA-Z]{1,3}$', self.email)
        if padrao:
            print('Email válido')
        else:
            print('Digite um email válido!')
            self.obter_email_usuario_senha()

    def raspagem_de_dados(self):
        chrome_options = Options()
        chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        chrome_options.add_argument('--lang=pr-BR')
        chrome_options.add_argument('--disable-notifications')

        # Usando o ChromeDriverManager para gerenciar o download do chromedriver
        self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        self.driver.set_window_size(800, 700)
        self.link = 'https://telefonesimportados.netlify.app/'
        self.lista_nome_celulares = []
        self.lista_preco_celulares = []
        self.driver.get(self.link)
        sleep(2)
        for p in range(5):
            item = 1
            for i in range(12):
                list_nomes = self.driver.find_elements(By.XPATH, f'/html/body/div[5]/div[2]/div[1]/div[{item}]/div/h2/a')
                self.lista_nome_celulares.append(list_nomes[0].text)
                sleep(1)
                lista_precos = self.driver.find_elements(By.XPATH, f'//div[{item}]/div[@class="single-shop-product" and 1]/div[@class="product-carousel-price" and 2]/ins[1]')
                self.lista_preco_celulares.append(lista_precos[0].text)
                item += 1
                sleep(1)
            try:
                botao_proximo = self.driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/div[2]/div/div/nav/ul/li[7]/a')
                botao_proximo.click()
                print(f'\u001b[32m{"Navegando para próxima página"}\u001b[0m')
                sleep(2)
            except NoSuchElementException:
                print(f'\u001b[32m{"Não há mais páginas!"}\u001b[0m')
                print(f'\u001b[32m{"Escaneamento Concluído"}\u001b[0m')
                self.driver.quit()

    def criar_planilha(self):
        index = 2
        planilha = openpyxl.Workbook()
        celulares = planilha['Sheet']
        celulares.title = 'Celulares'
        celulares['A1'] = 'Nome'
        celulares['B1'] = 'Preço'
        for nome, preco in zip(self.lista_nome_celulares, self.lista_preco_celulares):
            celulares.cell(column=1, row=index, value=nome)
            celulares.cell(column=2, row=index, value=preco)
            index += 1
        planilha.save("planilha_de_precos.xlsx")
        print(f'\u001b[32m{"Planilha criada com sucesso"}\u001b[0m')

    def enviar_email_cliente(self):
    endereco = 'seu-email@gmail.com'
    senha = 'senha-de-app-gerada-pelo-google'  # Insira a senha de app gerada
    msg = EmailMessage()
    msg['Subject'] = 'Planilha de Preços de Telefones Importados'
    msg['From'] = endereco
    msg['To'] = self.email
    msg.set_content('Olá, a sua planilha chegou.')
    arquivos = ["planilha_de_precos.xlsx"]
    for arquivo in arquivos:
        with open(arquivo, 'rb') as arq:
            dados = arq.read()
            nome_arquivo = arq.name
        msg.add_attachment(dados, maintype='application', subtype='octet-stream', filename=nome_arquivo)
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.ehlo()
    server.starttls()
    server.login(endereco, senha, initial_response_ok=True)
    server.send_message(msg)
    print(f'\u001b[32m{"Email enviado para destinatário"}\u001b[0m')
    server.quit()



# Executa o código
start = Scrappy()
start.iniciar()

